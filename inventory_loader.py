# -*- coding: utf-8 -*-
"""
엑셀 파일(domino_inventory_training.xlsx 구조)에서
재고·공급처·이메일 템플릿을 읽고, 발주 필요 항목을 분석합니다.
"""
import os
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import openpyxl

try:
    from urllib.request import urlopen
except ImportError:
    urlopen = None


# 시트별 컬럼 인덱스 (0-based, 헤더 다음 데이터 행 기준)
SUPPLIERS_HEADER = ["공급업체", "담당자", "이메일", "리드타임(일)", "품목"]
INVENTORY_HEADER = [
    "품목코드", "이름", "규격", "단위", "현재고", "안전재고", "MOQ",
    "공급업체", "담당자명", "공급업체이메일", "리드타임(일)",
    "발주수량", "발주기준수량", "상태", "발주메시지"
]


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def load_suppliers(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    result = []
    for row in rows[1:]:
        if not row or _cell_str(row[0]) == "":
            continue
        result.append({
            "name": _cell_str(row[0]),
            "contact": _cell_str(row[1]) if len(row) > 1 else "",
            "email": _cell_str(row[2]) if len(row) > 2 else "",
            "lead_time_days": int(row[3]) if len(row) > 3 and row[3] is not None else 0,
            "items": _cell_str(row[4]) if len(row) > 4 else "",
        })
    return result


def load_inventory(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    result = []
    for row in rows[1:]:
        if not row or _cell_str(row[0]) == "":
            continue
        try:
            current = int(row[4]) if row[4] is not None else 0
            safety = int(row[5]) if row[5] is not None else 0
            moq = int(row[6]) if row[6] is not None else 0
        except (TypeError, ValueError):
            current, safety, moq = 0, 0, 0
        # 발주수량 = MAX(MOQ, 안전재고 - 현재고), 최소 0
        order_qty = max(0, max(moq, safety - current))
        needs_order = order_qty > 0
        result.append({
            "code": _cell_str(row[0]),
            "name": _cell_str(row[1]) if len(row) > 1 else "",
            "spec": _cell_str(row[2]) if len(row) > 2 else "",
            "unit": _cell_str(row[3]) if len(row) > 3 else "",
            "current_stock": current,
            "safety_stock": safety,
            "moq": moq,
            "supplier": _cell_str(row[7]) if len(row) > 7 else "",
            "contact": _cell_str(row[8]) if len(row) > 8 else "",
            "supplier_email": _cell_str(row[9]) if len(row) > 9 else "",
            "lead_time_days": int(row[10]) if len(row) > 10 and row[10] is not None else 0,
            "order_quantity": order_qty,
            "status": "발주 필요" if needs_order else "정상",
            "order_message": _cell_str(row[14]) if len(row) > 14 else "",
        })
    return result


def load_email_template(ws) -> Dict[str, str]:
    subject = ""
    body = ""
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            val = _cell_str(cell)
            if "제목" in val and "형식" in val and i + 1 < len(row):
                subject = _cell_str(row[i + 1]) or subject
            if "본문" in val and "형식" in val and i + 1 < len(row):
                body = _cell_str(row[i + 1]) or body
    return {"subject": subject, "body": body}


def get_sheet_by_name(wb, name: str):
    for s in wb.worksheets:
        if s.title.strip() == name.strip():
            return s
    return None


def _find_inventory_sheet(wb):
    """이름으로 찾거나, 첫 행에 재고 관련 키워드가 있는 시트 반환."""
    inventory_ws = get_sheet_by_name(wb, "Inventory") or get_sheet_by_name(wb, "재고")
    if inventory_ws:
        return inventory_ws
    for ws in wb.worksheets:
        first_row = next(ws.iter_rows(values_only=True), None)
        if not first_row:
            continue
        first_str = " ".join(_cell_str(c) for c in first_row)
        if "품목코드" in first_str or "현재고" in first_str or "안전재고" in first_str:
            return ws
    return None


def load_all(excel_path: str) -> Dict[str, Any]:
    excel_path = (excel_path or "").strip()
    if not excel_path:
        return {"error": "파일 경로가 없습니다.", "suppliers": [], "inventory": [], "email_template": {}}
    err, suppliers, inventory, email_template = None, [], [], {"subject": "", "body": ""}
    if excel_path.startswith(("http://", "https://")) and urlopen:
        try:
            with urlopen(excel_path, timeout=30) as resp:
                data = resp.read()
            wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        except Exception as e:
            return {"error": f"URL에서 엑셀 로드 실패: {e}", "suppliers": [], "inventory": [], "email_template": {}}
    else:
        path = Path(excel_path)
        if not path.exists():
            return {"error": f"파일 없음: {excel_path}", "suppliers": [], "inventory": [], "email_template": {}}
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        suppliers_ws = get_sheet_by_name(wb, "Suppliers") or get_sheet_by_name(wb, "공급업체")
        inventory_ws = _find_inventory_sheet(wb)
        email_ws = get_sheet_by_name(wb, "EmailTemplate") or get_sheet_by_name(wb, "이메일템플릿")

        suppliers = load_suppliers(suppliers_ws) if suppliers_ws else []
        inventory = load_inventory(inventory_ws) if inventory_ws else []
        email_template = load_email_template(email_ws) if email_ws else {"subject": "", "body": ""}
        if not inventory_ws:
            err = "엑셀에서 재고 시트(Inventory 또는 '품목코드/현재고'가 있는 시트)를 찾을 수 없습니다."
    finally:
        wb.close()

    return {
        "error": err,
        "suppliers": suppliers,
        "inventory": inventory,
        "email_template": email_template,
    }


def update_inventory_item(
    excel_path: str,
    item_code: str,
    *,
    current_stock: Optional[int] = None,
    safety_stock: Optional[int] = None,
    moq: Optional[int] = None,
    name: Optional[str] = None,
    spec: Optional[str] = None,
    unit: Optional[str] = None,
    supplier: Optional[str] = None,
) -> Tuple[bool, str]:
    """
    엑셀 재고 시트에서 품목코드에 해당하는 행을 수정합니다.
    반환: (성공 여부, 메시지)
    엑셀 컬럼(1-based): A=품목코드, B=이름, C=규격, D=단위, E=현재고, F=안전재고, G=MOQ, H=공급업체
    """
    # Vercel 등 배포 환경: 디스크 읽기 전용이라 엑셀 저장 불가
    if os.environ.get("VERCEL") == "1":
        return False, "배포 환경(Vercel)에서는 엑셀 수정이 불가합니다. 수량·재고 변경은 로컬에서 실행할 때만 저장됩니다."
    path = Path(excel_path)
    if not path.exists():
        return False, "파일이 없습니다."
    if not item_code or not item_code.strip():
        return False, "품목코드를 입력하세요."
    wb = openpyxl.load_workbook(path, read_only=False)
    try:
        inv_ws = _find_inventory_sheet(wb)
        if not inv_ws:
            return False, "재고 시트를 찾을 수 없습니다."
        code_str = str(item_code).strip()
        data_row = None
        for row in range(2, inv_ws.max_row + 1):
            if _cell_str(inv_ws.cell(row=row, column=1).value) == code_str:
                data_row = row
                break
        if not data_row:
            return False, f"품목코드 '{code_str}'를 찾을 수 없습니다."
        if current_stock is not None:
            inv_ws.cell(row=data_row, column=5, value=int(current_stock))
        if safety_stock is not None:
            inv_ws.cell(row=data_row, column=6, value=int(safety_stock))
        if moq is not None:
            inv_ws.cell(row=data_row, column=7, value=int(moq))
        if name is not None:
            inv_ws.cell(row=data_row, column=2, value=str(name))
        if spec is not None:
            inv_ws.cell(row=data_row, column=3, value=str(spec))
        if unit is not None:
            inv_ws.cell(row=data_row, column=4, value=str(unit))
        if supplier is not None:
            inv_ws.cell(row=data_row, column=8, value=str(supplier))
        wb.save(path)
        return True, "저장되었습니다."
    except Exception as e:
        return False, str(e)
    finally:
        wb.close()


def add_inventory_item(
    excel_path: str,
    *,
    code: str,
    name: str = "",
    spec: str = "",
    unit: str = "",
    current_stock: int = 0,
    safety_stock: int = 0,
    moq: int = 0,
    supplier: str = "",
    contact: str = "",
    supplier_email: str = "",
    lead_time_days: int = 0,
) -> Tuple[bool, str]:
    """
    재고 시트에 새 품목 행을 추가합니다.
    반환: (성공 여부, 메시지)
    """
    if os.environ.get("VERCEL") == "1":
        return False, "배포 환경(Vercel)에서는 엑셀 추가가 불가합니다."
    path = Path(excel_path)
    if not path.exists():
        return False, "파일이 없습니다."
    code_str = (code or "").strip()
    if not code_str:
        return False, "품목코드를 입력하세요."
    wb = openpyxl.load_workbook(path, read_only=False)
    try:
        inv_ws = _find_inventory_sheet(wb)
        if not inv_ws:
            return False, "재고 시트를 찾을 수 없습니다."
        for row in range(2, inv_ws.max_row + 1):
            if _cell_str(inv_ws.cell(row=row, column=1).value) == code_str:
                return False, f"품목코드 '{code_str}'가 이미 존재합니다."
        next_row = inv_ws.max_row + 1
        inv_ws.cell(row=next_row, column=1, value=code_str)
        inv_ws.cell(row=next_row, column=2, value=str(name or ""))
        inv_ws.cell(row=next_row, column=3, value=str(spec or ""))
        inv_ws.cell(row=next_row, column=4, value=str(unit or ""))
        inv_ws.cell(row=next_row, column=5, value=int(current_stock))
        inv_ws.cell(row=next_row, column=6, value=int(safety_stock))
        inv_ws.cell(row=next_row, column=7, value=int(moq))
        inv_ws.cell(row=next_row, column=8, value=str(supplier or ""))
        inv_ws.cell(row=next_row, column=9, value=str(contact or ""))
        inv_ws.cell(row=next_row, column=10, value=str(supplier_email or ""))
        inv_ws.cell(row=next_row, column=11, value=int(lead_time_days))
        wb.save(path)
        return True, "항목이 추가되었습니다."
    except Exception as e:
        return False, str(e)
    finally:
        wb.close()


def get_orders_by_supplier(inventory: List[Dict]) -> List[Dict[str, Any]]:
    """발주 필요한 항목만 공급업체별로 묶습니다."""
    by_supplier: Dict[str, List[Dict]] = {}
    for item in inventory:
        if item["order_quantity"] <= 0:
            continue
        name = item["supplier"] or "(미지정)"
        if name not in by_supplier:
            by_supplier[name] = []
        by_supplier[name].append(item)
    result = []
    for name, items in by_supplier.items():
        first = items[0]
        total_qty = sum(i["order_quantity"] for i in items)
        result.append({
            "supplier_name": name,
            "contact": first.get("contact", ""),
            "email": first.get("supplier_email", ""),
            "lead_time_days": first.get("lead_time_days", 0),
            "item_count": len(items),
            "total_order_quantity": total_qty,
            "items": items,
        })
    return result
