# -*- coding: utf-8 -*-
"""
재고 파악 → 부족 시 담당 기업에 발주서 이메일 발송 웹 시스템.
데이터: 엑셀 파일(domino_inventory_training.xlsx 구조) 업로드 또는 기본 파일 사용.
"""
import json
import os
from pathlib import Path
from time import time
from datetime import datetime, timezone, timedelta
from io import BytesIO
import openpyxl

_base_dir = Path(__file__).resolve().parent
_env_file = _base_dir / ".env"


def _load_env():
    """ .env 파일을 읽어 os.environ에 반영 (dotenv 미설치/실패 시 대비) """
    if not _env_file.exists():
        return
    try:
        with open(_env_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    key, _, value = line.partition("=")
                    key = key.strip()
                    value = value.strip().strip('"').strip("'")
                    if key:
                        os.environ[key] = value
    except Exception:
        pass


try:
    from dotenv import load_dotenv
    load_dotenv(_env_file)
except ImportError:
    pass
_load_env()

from flask import Flask, request, render_template, jsonify, redirect, url_for, session, send_file
from werkzeug.utils import secure_filename

from inventory_loader import load_all, get_orders_by_supplier, update_inventory_item, add_inventory_item
from email_sender import fill_template, send_order_email, DEFAULT_SENDER_EMAIL, DEFAULT_STORE_NAME, DEFAULT_INTERNAL_OWNER

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "inventory-dev-secret-key")
# 세션 쿠키만 사용(브라우저 닫으면 만료) → 다시 들어올 때마다 로그인 필요
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10MB
# Vercel: 쓰기 가능한 경로 사용
if os.environ.get("VERCEL") == "1":
    app.config["UPLOAD_FOLDER"] = Path("/tmp/uploads")
else:
    app.config["UPLOAD_FOLDER"] = _base_dir / "uploads"
app.config["DEFAULT_EXCEL"] = _base_dir / "domino_inventory_training.xlsx"
try:
    app.config["UPLOAD_FOLDER"].mkdir(parents=True, exist_ok=True)
except OSError:
    pass

TEAM_PASSWORD = os.environ.get("TEAM_PASSWORD", "1234")

ALLOWED_EXTENSIONS = {"xlsx", "xls"}

# 해당 주소로는 발주 메일 발송 안 함 (소문자로 비교)
BLOCKED_EMAILS = {e.strip().lower() for e in os.environ.get("BLOCKED_EMAILS", "liszzm@naver.com").split(",") if e.strip()}

# 최근 발송일시 저장 (공급업체명 -> "YYYY-MM-DD HH:MM:SS"). 배포 환경은 읽기 전용일 수 있으므로 /tmp 사용.
if os.environ.get("VERCEL") == "1":
    _last_sent_dir = Path("/tmp")
else:
    _last_sent_dir = _base_dir / "data"
try:
    _last_sent_dir.mkdir(parents=True, exist_ok=True)
except OSError:
    _last_sent_dir = Path("/tmp")
    try:
        _last_sent_dir.mkdir(parents=True, exist_ok=True)
    except OSError:
        pass
LAST_SENT_FILE = _last_sent_dir / "last_sent.json"


def _load_last_sent() -> dict:
    """공급업체별 최근 발송일시 로드."""
    if not LAST_SENT_FILE.exists():
        return {}
    try:
        with open(LAST_SENT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


KST = timezone(timedelta(hours=9))


def _save_last_sent(supplier_name: str, dt: datetime):
    """해당 공급업체 최근 발송일시 갱신 (저장은 KST 문자열)."""
    data = _load_last_sent()
    data[supplier_name] = dt.strftime("%Y-%m-%d %H:%M:%S")
    try:
        with open(LAST_SENT_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=0)
    except Exception:
        pass


def apply_inventory_overrides(inventory_list):
    """세션에 저장된 수량 수정(배포 환경용)을 재고 목록에 반영. 발주수량 직접 수정값이 있으면 사용, 없으면 재계산."""
    if inventory_list is None:
        return []
    overrides = session.get("inventory_overrides") or {}
    if not overrides:
        return list(inventory_list)
    result = []
    for item in inventory_list:
        item = dict(item)
        code = (item.get("code") or "").strip()
        if code in overrides:
            o = overrides[code]
            if "current_stock" in o: item["current_stock"] = int(o["current_stock"])
            if "safety_stock" in o: item["safety_stock"] = int(o["safety_stock"])
            if "moq" in o: item["moq"] = int(o["moq"])
            if "order_quantity" in o:
                item["order_quantity"] = max(0, int(o["order_quantity"]))
                item["status"] = "발주 필요" if item["order_quantity"] > 0 else "정상"
                result.append(item)
                continue
        current = item.get("current_stock", 0)
        safety = item.get("safety_stock", 0)
        moq = item.get("moq", 0)
        order_qty = max(0, max(moq, safety - current))
        item["order_quantity"] = order_qty
        item["status"] = "발주 필요" if order_qty > 0 else "정상"
        result.append(item)
    return result


def _item_with_order_status(item):
    """item에 발주수량·상태 계산해 넣어서 반환."""
    d = dict(item)
    current = d.get("current_stock", 0)
    safety = d.get("safety_stock", 0)
    moq = d.get("moq", 0)
    order_qty = max(0, max(moq, safety - current))
    d["order_quantity"] = order_qty
    d["status"] = "발주 필요" if order_qty > 0 else "정상"
    return d


def get_effective_inventory(excel_path: str):
    """엑셀 + 세션 오버라이드 + 배포 환경 세션 추가 항목을 합친 재고 목록 반환. last_sent 포함."""
    try:
        data = load_all(excel_path)
    except Exception:
        return []
    if data.get("error"):
        return []
    inv = data.get("inventory")
    if inv is None:
        inv = []
    inventory = apply_inventory_overrides(inv)
    last_sent_map = _load_last_sent()
    for item in inventory:
        item["last_sent"] = last_sent_map.get((item.get("supplier") or "").strip(), "") or ""
    added = session.get("inventory_added_items")
    if not isinstance(added, list):
        added = []
    existing_codes = {(i.get("code") or "").strip() for i in inventory}
    for a in added:
        if not isinstance(a, dict):
            continue
        code = (a.get("code") or "").strip()
        if not code or code in existing_codes:
            continue
        existing_codes.add(code)
        item = _item_with_order_status(a)
        item["last_sent"] = last_sent_map.get((item.get("supplier") or "").strip(), "") or ""
        item.setdefault("contact", a.get("contact", ""))
        item.setdefault("supplier_email", a.get("supplier_email", ""))
        item.setdefault("lead_time_days", a.get("lead_time_days", 0))
        inventory.append(item)
    return inventory


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[-1].lower() in ALLOWED_EXTENSIONS


def auth_required():
    if session.get("auth") is True:
        return None
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if (request.form.get("password") or "").strip() == TEAM_PASSWORD:
            session["auth"] = True
            session.permanent = False  # 브라우저 닫으면 세션 만료, 재방문 시 로그인 필요
            return redirect(url_for("index"))
        return render_template("login.html", error="비밀번호가 올바르지 않습니다.")
    if session.get("auth") is True:
        return redirect(url_for("index"))
    return render_template("login.html", error=None)


@app.route("/logout")
def logout():
    session.pop("auth", None)
    return redirect(url_for("login"))


@app.errorhandler(500)
def handle_500(e):
    try:
        return render_template(
            "index.html",
            **_index_context(
                error="서버 오류가 발생했습니다. 잠시 후 다시 시도해 주세요.",
                read_only_deploy=os.environ.get("VERCEL") == "1",
            ),
        )
    except Exception:
        return "<h1>오류 발생</h1><p>잠시 후 다시 시도해 주세요.</p>", 500


@app.before_request
def require_auth():
    if request.endpoint in ("login", "static") or request.path.startswith("/static"):
        return None
    return auth_required()


def _index_context(error=None, orders=None, inventory=None, summary=None, chart_by_status=None, chart_by_supplier=None, excel_path=None, excel_filename=None, excel_path_for_api=None, email_template=None, read_only_deploy=False):
    """index.html에 넘길 공통 컨텍스트. None인 필드는 안전한 기본값. excel_path_for_api: API 호출 시 사용할 경로(업로드 파일이면 파일명만)."""
    fname = excel_filename if excel_filename is not None else (Path(excel_path).name if excel_path else "")
    upload_dir = Path(app.config["UPLOAD_FOLDER"])
    api_path = excel_path_for_api
    if api_path is None and excel_path:
        try:
            p = Path(excel_path).resolve()
            ud = upload_dir.resolve()
            if str(p).startswith(str(ud)) or p == ud:
                api_path = p.name
        except (ValueError, OSError):
            pass
        if api_path is None:
            api_path = excel_path
    return {
        "error": error,
        "orders": orders or [],
        "inventory": inventory or [],
        "summary": summary,
        "chart_by_status": chart_by_status if chart_by_status is not None else [],
        "chart_by_supplier": chart_by_supplier if chart_by_supplier is not None else [],
        "excel_path": excel_path or "",
        "excel_path_for_api": api_path or excel_path or "",
        "excel_filename": fname or "domino_inventory_training.xlsx",
        "email_template": email_template or {},
        "read_only_deploy": read_only_deploy,
    }


@app.route("/")
def index():
    try:
        file_param = request.args.get("file")
        if file_param and ".." not in file_param:
            session["uploaded_file"] = file_param
        # 업로드한 파일명(URL 또는 세션)을 우선 사용해 경로 해석
        excel_path = resolve_excel_path(file_param or session.get("uploaded_file") or "")
        # Blob URL이면 exists 검사 생략 (배포 환경 업로드 반영)
        if not (excel_path.startswith("http://") or excel_path.startswith("https://")):
            try:
                if not Path(excel_path).resolve().exists():
                    excel_path = str(Path(app.config["DEFAULT_EXCEL"]).resolve())
                    if not Path(excel_path).exists():
                        session.pop("uploaded_file", None)
                        session.pop("uploaded_blob_url", None)
            except (OSError, RuntimeError):
                excel_path = str(app.config["DEFAULT_EXCEL"])
                session.pop("uploaded_file", None)
                session.pop("uploaded_blob_url", None)

        data = load_all(excel_path)
        if data.get("error"):
            return render_template(
                "index.html",
                **_index_context(
                    error=data["error"],
                    excel_path=excel_path,
                    email_template=data.get("email_template"),
                    read_only_deploy=os.environ.get("VERCEL") == "1",
                ),
            )

        inventory = get_effective_inventory(excel_path)
        orders = get_orders_by_supplier(inventory)
        need_count = sum(1 for i in inventory if i.get("order_quantity", 0) > 0)
        total_items = len(inventory)
        total_order_qty = sum(i.get("order_quantity", 0) for i in inventory)
        summary = {
            "total_items": total_items,
            "need_order_count": need_count,
            "total_order_quantity": total_order_qty,
            "supplier_count": len(orders),
        }
        chart_by_status = [
            {"label": "정상", "count": total_items - need_count, "pct": round(100 * (total_items - need_count) / total_items, 1) if total_items else 0},
            {"label": "발주 필요", "count": need_count, "pct": round(100 * need_count / total_items, 1) if total_items else 0},
        ]
        max_supplier_qty = max((o.get("total_order_quantity") or 0 for o in orders), default=1)
        chart_by_supplier = [
            {"name": o.get("supplier_name") or "(미지정)", "qty": o.get("total_order_quantity") or 0, "pct": round(100 * (o.get("total_order_quantity") or 0) / max_supplier_qty, 1) if max_supplier_qty else 0}
            for o in orders
        ]
        return render_template(
            "index.html",
            **_index_context(
                orders=orders,
                inventory=inventory,
                summary=summary,
                chart_by_status=chart_by_status,
                chart_by_supplier=chart_by_supplier,
                excel_path=excel_path,
                email_template=data.get("email_template"),
                read_only_deploy=os.environ.get("VERCEL") == "1",
            ),
        )
    except Exception as e:
        err_msg = str(e)
        try:
            return render_template(
                "index.html",
                **_index_context(
                    error=f"오류가 발생했습니다: {err_msg}",
                    excel_path=excel_path,
                    read_only_deploy=os.environ.get("VERCEL") == "1",
                ),
            )
        except Exception:
            return (
                "<!DOCTYPE html><html><head><meta charset='utf-8'><title>오류</title></head><body>"
                "<h1>오류 발생</h1><p>잠시 후 다시 시도해 주세요.</p>"
                "<pre style='background:#f5f5f5;padding:1rem;font-size:0.875rem;'>" + err_msg.replace("<", "&lt;") + "</pre>"
                "</body></html>",
                200,
            )


def _upload_to_vercel_blob(safe_name: str, file_bytes: bytes):
    """배포 환경에서 업로드 파일을 Vercel Blob에 저장하고 URL 반환. 실패 시 None. (vercel-blob 패키지 및 BLOB_READ_WRITE_TOKEN 필요)"""
    if not os.environ.get("BLOB_READ_WRITE_TOKEN"):
        return None
    try:
        import vercel_blob  # optional: pip install vercel-blob
        resp = vercel_blob.put(safe_name, file_bytes)
        if resp and hasattr(resp, "url"):
            return getattr(resp, "url", None)
        if isinstance(resp, dict):
            return resp.get("url") or resp.get("downloadUrl")
        return None
    except Exception:
        return None


@app.route("/upload", methods=["POST"])
def upload():
    if "excel" not in request.files:
        return redirect(url_for("index"))
    f = request.files["excel"]
    if f.filename == "" or not allowed_file(f.filename):
        return redirect(url_for("index"))
    upload_dir = Path(app.config["UPLOAD_FOLDER"])
    upload_dir.mkdir(parents=True, exist_ok=True)
    ext = (f.filename or "").rsplit(".", 1)[-1].lower() if "." in (f.filename or "") else "xlsx"
    if ext not in ALLOWED_EXTENSIONS:
        ext = "xlsx"
    safe_name = secure_filename(f.filename) or f"upload_{int(time())}.{ext}"
    if not safe_name.strip():
        safe_name = f"upload_{int(time())}.{ext}"
    path = upload_dir / safe_name
    file_bytes = f.read()
    try:
        path.write_bytes(file_bytes)
    except (PermissionError, OSError) as e:
        return render_template(
            "index.html",
            error=f"파일 저장 실패(권한 오류): {e}. 'uploads' 폴더 쓰기 권한을 확인하세요.",
            orders=[],
            inventory=[],
            summary=None,
            chart_by_status=[],
            chart_by_supplier=[],
            excel_path=str(app.config["DEFAULT_EXCEL"]),
            email_template={},
            read_only_deploy=os.environ.get("VERCEL") == "1",
        )
    session["uploaded_file"] = safe_name
    session["uploaded_blob_url"] = None
    # 새 엑셀 업로드 시 웹에서 수정한 값(세션 오버라이드) 초기화 → 업로드한 엑셀 내용 그대로 반영
    session.pop("inventory_overrides", None)
    session.pop("inventory_added_items", None)
    if os.environ.get("VERCEL") == "1":
        blob_url = _upload_to_vercel_blob(safe_name, file_bytes)
        if blob_url:
            session["uploaded_blob_url"] = blob_url

    # JSON 요청이면 업로드 직후 파싱 데이터를 반환 (리다이렉트 없이 화면 반영용)
    if request.accept_mimetypes.best_match(["application/json", "text/html"]) == "application/json":
        try:
            data = load_all(str(path))
            if data.get("error"):
                return jsonify({"ok": False, "message": data["error"]}), 400
            inventory = get_effective_inventory(str(path))
            orders = get_orders_by_supplier(inventory)
            need_count = sum(1 for i in inventory if i.get("order_quantity", 0) > 0)
            total_items = len(inventory)
            total_order_qty = sum(i.get("order_quantity", 0) for i in inventory)
            summary = {
                "total_items": total_items,
                "need_order_count": need_count,
                "total_order_quantity": total_order_qty,
                "supplier_count": len(orders),
            }
            chart_by_status = [
                {"label": "정상", "count": total_items - need_count, "pct": round(100 * (total_items - need_count) / total_items, 1) if total_items else 0},
                {"label": "발주 필요", "count": need_count, "pct": round(100 * need_count / total_items, 1) if total_items else 0},
            ]
            max_supplier_qty = max((o.get("total_order_quantity") or 0 for o in orders), default=1)
            chart_by_supplier = [
                {"name": o.get("supplier_name") or "(미지정)", "qty": o.get("total_order_quantity") or 0, "pct": round(100 * (o.get("total_order_quantity") or 0) / max_supplier_qty, 1) if max_supplier_qty else 0}
                for o in orders
            ]
            return jsonify({
                "ok": True,
                "message": "업로드되었습니다. 재고현황에 반영되었습니다.",
                "excel_filename": safe_name,
                "excel_path_for_api": safe_name,
                "inventory": inventory,
                "orders": orders,
                "summary": summary,
                "chart_by_status": chart_by_status,
                "chart_by_supplier": chart_by_supplier,
            })
        except Exception as e:
            return jsonify({"ok": False, "message": str(e)}), 500

    return redirect(url_for("index", file=safe_name))


def resolve_excel_path(client_path: str) -> str:
    """클라이언트가 보낸 경로(또는 파일명)를 실제 사용할 엑셀 경로로 통일. 배포 시 Blob URL 우선."""
    # 배포 환경: 세션에 Blob URL이 있으면 그대로 사용 (업로드 파일 영구 반영)
    if os.environ.get("VERCEL") == "1" and session.get("uploaded_blob_url"):
        return session["uploaded_blob_url"]
    raw = (client_path or "").strip()
    if not raw:
        raw = session.get("uploaded_file") or ""
    if raw and ".." in raw:
        raw = ""
    upload_dir = Path(app.config["UPLOAD_FOLDER"]).resolve()
    default_path = Path(app.config["DEFAULT_EXCEL"]).resolve()
    default = str(default_path)
    p = Path(raw) if raw else None
    if p and p.is_absolute():
        try:
            resolved = p.resolve()
            if resolved.exists():
                return str(resolved)
        except (OSError, RuntimeError):
            pass
    if raw:
        name = Path(raw).name
        cand = (upload_dir / name).resolve()
        if cand.exists():
            return str(cand)
    if session.get("uploaded_file"):
        cand = (upload_dir / session["uploaded_file"]).resolve()
        if cand.exists():
            return str(cand)
    return default


@app.route("/api/inventory/update", methods=["POST"])
def api_inventory_update():
    """재고 품목 한 건 수정. 엑셀 파일에 반영. 성공 시 갱신된 항목·요약·시각화 반환. 업로드 파일 경로 우선."""
    data = request.get_json() or {}
    excel_path = resolve_excel_path(str(data.get("excel_path") or "").strip())
    item_code = data.get("item_code")
    if not item_code:
        return jsonify({"ok": False, "message": "품목코드가 필요합니다."})
    try:
        current_stock = data.get("current_stock")
        safety_stock = data.get("safety_stock")
        moq = data.get("moq")
        order_quantity = data.get("order_quantity")
        if current_stock is not None:
            current_stock = int(current_stock)
        if safety_stock is not None:
            safety_stock = int(safety_stock)
        if moq is not None:
            moq = int(moq)
        if order_quantity is not None:
            order_quantity = max(0, int(order_quantity))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "message": "현재고/안전재고/MOQ/발주수량은 숫자로 입력하세요."})
    code = str(item_code).strip()
    # 배포 환경(Vercel): 엑셀 저장 불가 → 세션에만 저장.
    if os.environ.get("VERCEL") == "1":
        overrides = session.get("inventory_overrides") or {}
        base = {k: v for k, v in ({"current_stock": current_stock, "safety_stock": safety_stock, "moq": moq, "order_quantity": order_quantity}).items() if v is not None}
        overrides[code] = {**(overrides.get(code) or {}), **base}
        session["inventory_overrides"] = overrides
        session.modified = True
    else:
        if not Path(excel_path).exists():
            return jsonify({"ok": False, "message": "엑셀 파일을 찾을 수 없습니다."})
        ok, msg = update_inventory_item(
            excel_path,
            code,
            current_stock=current_stock, safety_stock=safety_stock, moq=moq,
            order_quantity=order_quantity,
            name=data.get("name"), spec=data.get("spec"), unit=data.get("unit"), supplier=data.get("supplier"),
        )
        if not ok:
            return jsonify({"ok": False, "message": msg})
    # 갱신된 재고로 요약·시각화 계산해 바로 반영용으로 반환
    resolved = resolve_excel_path(excel_path)
    inventory = get_effective_inventory(resolved)
    orders = get_orders_by_supplier(inventory)
    need_count = sum(1 for i in inventory if i.get("order_quantity", 0) > 0)
    total_items = len(inventory)
    total_order_qty = sum(i.get("order_quantity", 0) for i in inventory)
    summary = {
        "total_items": total_items,
        "need_order_count": need_count,
        "total_order_quantity": total_order_qty,
        "supplier_count": len(orders),
    }
    chart_by_status = [
        {"label": "정상", "count": total_items - need_count, "pct": round(100 * (total_items - need_count) / total_items, 1) if total_items else 0},
        {"label": "발주 필요", "count": need_count, "pct": round(100 * need_count / total_items, 1) if total_items else 0},
    ]
    max_supplier_qty = max((o.get("total_order_quantity") or 0 for o in orders), default=1)
    chart_by_supplier = [
        {"name": o.get("supplier_name") or "(미지정)", "qty": o.get("total_order_quantity") or 0, "pct": round(100 * (o.get("total_order_quantity") or 0) / max_supplier_qty, 1) if max_supplier_qty else 0}
        for o in orders
    ]
    updated_item = next((i for i in inventory if (i.get("code") or "").strip() == code), None)
    payload = {
        "ok": True,
        "message": "저장되었습니다. 메일 발송 시 반영됩니다." if os.environ.get("VERCEL") == "1" else "저장되었습니다.",
        "item": {"code": code, "order_quantity": updated_item.get("order_quantity", 0), "status": updated_item.get("status", "정상")} if updated_item else None,
        "summary": summary,
        "chart_by_status": chart_by_status,
        "chart_by_supplier": chart_by_supplier,
    }
    return jsonify(payload)


@app.route("/api/inventory/add", methods=["POST"])
def api_inventory_add():
    """재고 품목 한 건 추가. 로컬은 엑셀에 저장, 배포(Vercel)는 세션에 저장해 화면·발송·엑셀내보내기에 반영."""
    data = request.get_json() or {}
    excel_path = resolve_excel_path(str(data.get("excel_path") or "").strip())
    if os.environ.get("VERCEL") != "1" and not Path(excel_path).exists():
        return jsonify({"ok": False, "message": "엑셀 파일을 찾을 수 없습니다."})
    code = (data.get("code") or "").strip()
    if not code:
        return jsonify({"ok": False, "message": "품목코드는 필수입니다."})
    try:
        current_stock = int(data.get("current_stock", 0))
        safety_stock = int(data.get("safety_stock", 0))
        moq = int(data.get("moq", 0))
        lead_time_days = int(data.get("lead_time_days", 0))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "message": "현재고/안전재고/MOQ/리드타임은 숫자로 입력하세요."})
    if os.environ.get("VERCEL") == "1":
        base_data = load_all(excel_path)
        existing_codes = {(i.get("code") or "").strip() for i in (base_data.get("inventory") or [])}
        added = session.get("inventory_added_items") or []
        for a in added:
            existing_codes.add((a.get("code") or "").strip())
        if code in existing_codes:
            return jsonify({"ok": False, "message": f"품목코드 '{code}'가 이미 존재합니다."})
        new_item = {
            "code": code,
            "name": (data.get("name") or "").strip(),
            "spec": (data.get("spec") or "").strip(),
            "unit": (data.get("unit") or "").strip(),
            "current_stock": current_stock,
            "safety_stock": safety_stock,
            "moq": moq,
            "supplier": (data.get("supplier") or "").strip(),
            "contact": (data.get("contact") or "").strip(),
            "supplier_email": (data.get("supplier_email") or "").strip(),
            "lead_time_days": lead_time_days,
        }
        session.setdefault("inventory_added_items", []).append(new_item)
        session.modified = True
        return jsonify({"ok": True, "message": "항목이 추가되었습니다. 화면·메일 발송·엑셀 내보내기에 반영됩니다."})
    ok, msg = add_inventory_item(
        excel_path,
        code=code,
        name=(data.get("name") or "").strip(),
        spec=(data.get("spec") or "").strip(),
        unit=(data.get("unit") or "").strip(),
        current_stock=current_stock,
        safety_stock=safety_stock,
        moq=moq,
        supplier=(data.get("supplier") or "").strip(),
        contact=(data.get("contact") or "").strip(),
        supplier_email=(data.get("supplier_email") or "").strip(),
        lead_time_days=lead_time_days,
    )
    return jsonify({"ok": ok, "message": msg})


def _build_export_inventory(excel_path: str):
    """현재 재고(오버라이드+세션 추가 항목+최근발송일시)로 엑셀 바이트 생성."""
    inventory = get_effective_inventory(excel_path)
    if not inventory and load_all(excel_path).get("error"):
        return None, load_all(excel_path)["error"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "재고현황"
    headers = [
        "품목코드", "이름", "규격", "단위", "현재고", "안전재고", "MOQ",
        "공급업체", "담당자명", "공급업체이메일", "리드타임(일)",
        "발주수량", "상태", "최근 발송일시"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, item in enumerate(inventory, 2):
        ws.cell(row=row_idx, column=1, value=item.get("code", ""))
        ws.cell(row=row_idx, column=2, value=item.get("name", ""))
        ws.cell(row=row_idx, column=3, value=item.get("spec", ""))
        ws.cell(row=row_idx, column=4, value=item.get("unit", ""))
        ws.cell(row=row_idx, column=5, value=item.get("current_stock", 0))
        ws.cell(row=row_idx, column=6, value=item.get("safety_stock", 0))
        ws.cell(row=row_idx, column=7, value=item.get("moq", 0))
        ws.cell(row=row_idx, column=8, value=item.get("supplier", ""))
        ws.cell(row=row_idx, column=9, value=item.get("contact", ""))
        ws.cell(row=row_idx, column=10, value=item.get("supplier_email", ""))
        ws.cell(row=row_idx, column=11, value=item.get("lead_time_days", 0))
        ws.cell(row=row_idx, column=12, value=item.get("order_quantity", 0))
        ws.cell(row=row_idx, column=13, value=item.get("status", ""))
        ws.cell(row=row_idx, column=14, value=item.get("last_sent", ""))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, None


def _excel_path_exists(excel_path: str) -> bool:
    """경로가 로컬 파일이면 존재 여부 반환, URL이면 True(사용 가능)."""
    if not excel_path or excel_path.startswith("http://") or excel_path.startswith("https://"):
        return bool(excel_path)
    try:
        return Path(excel_path).resolve().exists()
    except (OSError, RuntimeError):
        return False


@app.route("/api/inventory/export", methods=["POST"])
def api_inventory_export():
    """재고현황(최근 발송일시 포함) 엑셀 다운로드. 업로드한 파일 경로 우선 반영."""
    raw = (request.get_json() or {}).get("excel_path") or ""
    excel_path = resolve_excel_path(str(raw).strip())
    if not _excel_path_exists(excel_path):
        return jsonify({"ok": False, "message": "엑셀 파일을 찾을 수 없습니다."}), 400
    buf, err = _build_export_inventory(excel_path)
    if err:
        return jsonify({"ok": False, "message": err}), 400
    filename = f"재고현황_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/send-orders", methods=["POST"])
def api_send_orders():
    """선택한 공급업체에 대해 발주서 이메일 발송. 업로드한 파일 경로 우선 반영."""
    excel_path = resolve_excel_path(str((request.json or {}).get("excel_path") or "").strip())
    supplier_indices = request.json.get("supplier_indices")  # 보낼 공급업체 인덱스 목록 (전부면 생략 가능)
    store_name = request.json.get("store_name") or DEFAULT_STORE_NAME
    internal_owner = request.json.get("internal_owner") or DEFAULT_INTERNAL_OWNER

    sender_email = os.environ.get("SMTP_USER", DEFAULT_SENDER_EMAIL)
    sender_password = os.environ.get("SMTP_PASSWORD", "")
    smtp_host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))

    data = load_all(excel_path)
    if data.get("error"):
        return jsonify({"ok": False, "message": data["error"]})
    inventory = get_effective_inventory(excel_path)
    orders = get_orders_by_supplier(inventory)
    supplier_email_map = {}
    for s in data.get("suppliers", []):
        name = (s.get("name") or "").strip()
        email = (s.get("email") or "").strip()
        if name and email and "@" in email:
            supplier_email_map[name] = email
    for order in orders:
        if not order.get("email") or "@" not in str(order.get("email", "")):
            fallback = supplier_email_map.get((order.get("supplier_name") or "").strip())
            if fallback:
                order["email"] = fallback
    tpl = data.get("email_template", {})
    subject_tpl = tpl.get("subject") or "[발주요청] {{STORE_NAME}} / {{SUPPLIER_NAME}} / {{ORDER_DATE}}"
    body_tpl = tpl.get("body") or "안녕하세요 {{SUPPLIER_NAME}} 담당자님.\n\n{{STORE_NAME}}입니다.\n아래 품목 발주 요청 드립니다.\n\n{{ITEM_LIST}}\n\n확인 부탁드립니다.\n{{INTERNAL_OWNER}}"

    results = []
    for idx, order in enumerate(orders):
        if supplier_indices is not None and idx not in supplier_indices:
            continue
        subject, body = fill_template(
            subject_tpl, body_tpl,
            order["supplier_name"], order["items"],
            store_name=store_name, internal_owner=internal_owner,
        )
        to_email = (order.get("email") or "").strip().lower()
        if to_email in BLOCKED_EMAILS:
            ok, msg = False, "해당 주소는 발송 제외됩니다."
        else:
            ok, msg = send_order_email(
                order["email"], subject, body,
                sender_email=sender_email, sender_password=sender_password,
                smtp_host=smtp_host, smtp_port=smtp_port,
                bcc=os.environ.get("SMTP_BCC") or sender_email,
            )
        now_kst = datetime.now(KST)
        if ok:
            _save_last_sent(order["supplier_name"], now_kst)
        result = {
            "supplier": order["supplier_name"],
            "email": order["email"],
            "ok": ok,
            "message": msg,
        }
        if ok:
            result["last_sent"] = now_kst.strftime("%Y-%m-%d %H:%M:%S")
        results.append(result)
    return jsonify({"ok": True, "results": results})


if __name__ == "__main__":
    # 코드·.env 수정 시 서버 자동 재시작 (use_reloader=True)
    _extra = [str(_env_file)] if _env_file.exists() else []
    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True,
        use_reloader=True,
        extra_files=_extra,
    )
