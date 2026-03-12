# -*- coding: utf-8 -*-
"""엑셀 업로드 후 index에서 업로드한 파일이 반영되는지 테스트."""
import os
import sys
from pathlib import Path
from io import BytesIO

# 프로젝트 루트를 path에 추가
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl


def make_test_excel():
    """테스트용 최소 엑셀 생성 (Inventory 시트, 헤더+1행)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    headers = [
        "품목코드", "이름", "규격", "단위", "현재고", "안전재고", "MOQ",
        "공급업체", "담당자명", "공급업체이메일", "리드타임(일)",
        "발주수량", "발주기준수량", "상태", "발주메시지",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    # 고유한 품목코드로 업로드 반영 여부 확인
    ws.cell(row=2, column=1, value="TEST_UPLOAD_001")
    ws.cell(row=2, column=2, value="업로드테스트품목")
    ws.cell(row=2, column=3, value="")
    ws.cell(row=2, column=4, value="EA")
    ws.cell(row=2, column=5, value=10)
    ws.cell(row=2, column=6, value=20)
    ws.cell(row=2, column=7, value=5)
    ws.cell(row=2, column=8, value="테스트공급처")
    for col in range(9, 16):
        ws.cell(row=2, column=col, value=0 if col == 12 else "")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_upload_reflected():
    os.environ["TEAM_PASSWORD"] = "1234"
    # DEFAULT_EXCEL이 없을 수 있으므로 Vercel 플래그 없이 로컬 설정
    if "VERCEL" in os.environ:
        del os.environ["VERCEL"]
    from app import app
    upload_dir = Path(app.config["UPLOAD_FOLDER"]).resolve()
    upload_dir.mkdir(parents=True, exist_ok=True)
    test_filename = "test_upload_reflect_001.xlsx"
    excel_bytes = make_test_excel()
    with app.test_client() as client:
        # 로그인
        r = client.post("/login", data={"password": "1234"}, follow_redirects=True)
        if r.status_code != 200:
            raise AssertionError(f"Login failed: {r.status_code}")
        # 업로드
        excel_bytes.seek(0)
        r = client.post(
            "/upload",
            data={"excel": (excel_bytes, test_filename)},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        if r.status_code != 302:
            raise AssertionError(f"Upload expected 302, got {r.status_code}. Body: {r.data[:500]}")
        redirect_url = r.headers.get("Location", "")
        if "file=" not in redirect_url:
            raise AssertionError(f"Redirect should contain file= param: {redirect_url}")
        # 리다이렉트 따라가서 index 확인
        r2 = client.get(redirect_url)
        if r2.status_code != 200:
            raise AssertionError(f"Index after redirect: {r2.status_code}")
        html = r2.data.decode("utf-8", errors="replace")
        if "TEST_UPLOAD_001" not in html:
            raise AssertionError("Uploaded file data not reflected: TEST_UPLOAD_001 not in index HTML")
        if test_filename not in html and "현재 적용 파일" in html:
            # 현재 적용 파일 뒤에 파일명이 나와야 함
            if "test_upload_reflect" not in html:
                raise AssertionError("Uploaded filename not shown in '현재 적용 파일'")
    # 업로드된 파일이 실제로 저장되었는지 확인
    saved = upload_dir / test_filename
    if not saved.exists():
        raise AssertionError(f"Uploaded file not found on disk: {saved}")
    print("OK: Upload reflected in index and file on disk.")
    return True


def test_alternate_header_format():
    """재료명·현재재고·거래처·발주권장수량 등 다른 헤더 형식이 인식되는지 확인."""
    sys.path.insert(0, str(ROOT))
    from inventory_loader import load_all
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # 이미지와 같은 컬럼명 사용
    headers = [
        "품목코드", "재료명", "규격", "단위", "현재재고", "안전재고", "MOQ",
        "거래처", "알림담당자", "거래처이메일", "리드타임(일)",
        "부족수량", "발주권장수량", "상태",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=2, column=1, value="ING001")
    ws.cell(row=2, column=2, value="도우볼")
    ws.cell(row=2, column=5, value=120)
    ws.cell(row=2, column=6, value=180)
    ws.cell(row=2, column=7, value=100)
    ws.cell(row=2, column=8, value="도미노푸드서플라이")
    ws.cell(row=2, column=13, value=100)  # 발주권장수량
    ws.cell(row=2, column=14, value="발주 필요")
    path = ROOT / "uploads" / "test_alt_headers.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    try:
        data = load_all(str(path))
        assert not data.get("error"), data.get("error")
        inv = data.get("inventory") or []
        assert len(inv) >= 1, "inventory should have at least one row"
        first = inv[0]
        assert first.get("code") == "ING001"
        assert "도우볼" in first.get("name", "")
        assert first.get("current_stock") == 120
        assert first.get("safety_stock") == 180
        assert first.get("supplier", "").strip() != ""
        print("OK: Alternate header format (재료명/현재재고/거래처/발주권장수량) recognized.")
    finally:
        if path.exists():
            path.unlink(missing_ok=True)


if __name__ == "__main__":
    try:
        test_upload_reflected()
        test_alternate_header_format()
        print("All checks passed.")
    except Exception as e:
        print(f"FAIL: {e}")
        sys.exit(1)
